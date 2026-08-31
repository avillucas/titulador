import re
import openpyxl
from typing import List, Dict, Any, Optional
from src.models import Egresado, TituloData
from src.catalog import TrayectoCatalog, format_trayecto_data
from src.date_utils import calculate_default_emision_date


def format_dni(doc_val: Any) -> str:
    """Formats raw document number into dot-separated string (e.g. 35.140.353)."""
    if doc_val is None:
        return ""
    # Convert float to int if needed
    if isinstance(doc_val, float):
        doc_str = str(int(doc_val))
    else:
        doc_str = str(doc_val).strip()
    
    # Remove existing dots/spaces
    clean_digits = re.sub(r'\D', '', doc_str)
    if not clean_digits:
        return doc_str
    
    # Format with dots
    try:
        return f"{int(clean_digits):,}".replace(",", ".")
    except ValueError:
        return doc_str

def parse_fecha_egreso(text: str) -> str:
    """Extracts formatted date (e.g. '14 de Julio de 2025') from the header string in cell A7."""
    if not text:
        return ""
    # Example text: "... a los…14... dias del mes de julio del año  2025 se reúne..."
    match = re.search(r'los\s*[…\.]*\s*(\d{1,2})\s*[…\.]*\s*d[ií]as\s+del\s+mes\s+de\s+([a-zA-ZáéíóúñÁÉÍÓÚÑ]+)\s+del\s+a[ñn]o\s+(\d{4})', text, re.IGNORECASE)
    if match:
        day, month, year = match.groups()
        return f"{day.zfill(2)} de {month.capitalize()} de {year}"
    return text

class ExcelParser:
    def __init__(self, file_path: str):
        self.file_path = file_path

    def load_data(self, catalog: Optional[TrayectoCatalog] = None) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        sheet = wb.active if "Acta de examen" not in wb.sheetnames else wb["Acta de examen"]
        
        cpf_num = str(sheet.cell(3, 3).value or "412").strip()
        distrito = str(sheet.cell(3, 7).value or "Lomas de Zamora").strip()
        
        # Especialidad
        especialidad = str(sheet.cell(7, 13).value or "").strip()
        
        # Fecha egreso string in cell A7
        header_text = str(sheet.cell(7, 1).value or "")
        fecha_egreso = parse_fecha_egreso(header_text)
        emision_dia, emision_mes, emision_ano = calculate_default_emision_date(fecha_egreso)
        
        egresados: List[Dict[str, Any]] = []
        omitidos: List[Dict[str, Any]] = []
        
        # Scan rows starting from row 11 up to summary rows
        for r in range(11, sheet.max_row + 1):
            num_egresado_raw = sheet.cell(r, 1).value
            nombre = sheet.cell(r, 4).value
            doc_raw = sheet.cell(r, 20).value
            
            if not nombre:
                continue

            nombre_str = str(nombre).strip()
            # Ignore signature lines, footers, or headers
            if nombre_str.startswith(".") or nombre_str.startswith("Vocal") or "Comisión" in nombre_str or "Examinados" in nombre_str:
                continue
                
            num_egresado = ""
            if num_egresado_raw is not None:
                val_str = str(num_egresado_raw).strip()
                if val_str and not re.match(r'^-+$', val_str):
                    if isinstance(num_egresado_raw, float):
                        num_egresado = str(int(num_egresado_raw))
                    else:
                        num_egresado = val_str
            
            doc_formatted = format_dni(doc_raw)
            
            if num_egresado:
                egresados.append({
                    "num_egresado": num_egresado,
                    "apellido_nombre": nombre_str,
                    "documento": doc_formatted,
                    "estado": "Aprobado",
                    "fila": r
                })
            else:
                omitidos.append({
                    "num_egresado": "",
                    "apellido_nombre": nombre_str,
                    "documento": doc_formatted,
                    "estado": "Ausente/Desaprobado",
                    "motivo": "Sin número de egresado (no aprobó)",
                    "fila": r
                })

        if catalog is None:
            catalog = TrayectoCatalog()

        matched_trayecto = catalog.search_trayecto(especialidad)
        trayecto_data = format_trayecto_data(matched_trayecto) if matched_trayecto else None
            
        return {
            "numero_cpf": cpf_num,
            "distrito_cpf": distrito,
            "especialidad": especialidad,
            "fecha_egreso": fecha_egreso,
            "emision_dia": emision_dia,
            "emision_mes": emision_mes,
            "emision_ano": emision_ano,
            "egresados": egresados,
            "omitidos": omitidos,
            "trayecto_matched": matched_trayecto,
            "trayecto_data": trayecto_data
        }


